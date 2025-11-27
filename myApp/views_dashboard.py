# myApp/views_dashboard.py
from functools import wraps
from decimal import Decimal

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login as dj_login, logout as dj_logout
from django.db.models import Sum, Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from django.template.loader import render_to_string
from django.utils.html import strip_tags  # optional; handy if you add text fallbacks

from .models import Order
# already present:
from django.db.models import Q, Case, When, IntegerField
from django.utils.dateparse import parse_date

# add Product import (you have this model)
from .models import Product

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

# PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False


# Use the same mail shim & money formatter you wired for Resend
from .views import _safe_send_mail
try:
    from .views import money_filter as _money  # mirrors your template money filter
except Exception:
    def _money(val, request=None):
        try:
            return f"{Decimal(str(val or '0')):,.2f}"
        except Exception:
            return "0.00"
        
def _order_items_rel_name():
    return 'items' if any(f.name == 'items' for f in Order._meta.get_fields()) else 'orderitem_set'

def _item_has_product_fk(item_model):
    try:
        f = item_model._meta.get_field('product')
        return getattr(f, 'is_relation', False)
    except Exception:
        return False

def _model_has_field(model, field_name):
    try:
        model._meta.get_field(field_name)
        return True
    except Exception:
        return False
    

# --- add this helper near the top with your other helpers ---
# views_dashboard.py (near the top, after imports)
# views_dashboard.py (near your other helpers)

def _shipping_text_for(order) -> str:
    """
    Returns a nice multi-line shipping string from:
      1) order.shipping_address_text (if present)
      2) order.shipping_address (JSON)
      3) legacy flat fields (address_line1/city/province/zip_code)
    """
    # 1) nicest: what checkout already saved
    txt = (getattr(order, "shipping_address_text", "") or "").strip()
    if txt:
        return txt

    # 2) build from normalized JSON
    data = getattr(order, "shipping_address", None) or {}
    if isinstance(data, dict) and data:
        parts = []

        def add(val):
            v = (val or "").strip()
            if v:
                parts.append(v)

        # street lines
        add(data.get("address_line1"))
        add(data.get("address_line2"))

        # locality cluster (PH/JO/US/AE… tolerant)
        locality_bits = [
            data.get("barangay"),
            data.get("city"),
            # for JO allow either chosen option or free text
            data.get("area_other") or data.get("area"),
            data.get("province") or data.get("state") or data.get("emirate") or data.get("county"),
        ]
        add(", ".join([b.strip() for b in locality_bits if (b or "").strip()]))

        # postal
        add(data.get("postal_code") or data.get("zip_code"))

        # country (ISO-2)
        add((getattr(order, "country", "") or "").upper())

        return "\n".join([p for p in parts if p]).strip()

    # 3) fallback to legacy fields
    legacy_parts = []
    for v in [
        getattr(order, "address_line1", ""),
        ", ".join(
            [x for x in [
                getattr(order, "city", ""),
                getattr(order, "province", "")
            ] if (x or "").strip()]
        ),
        getattr(order, "zip_code", ""),
        (getattr(order, "country", "") or "").upper(),
    ]:
        v = (v or "").strip()
        if v:
            legacy_parts.append(v)

    return "\n".join(legacy_parts).strip()




def _shipping_address_text(order):
    """
    Best-effort pretty address string.
    Prefers order.shipping_address_text, then order.shipping_address (JSON),
    and finally the legacy flat fields (address_line1/city/province/zip_code).
    """
    # 1) explicit formatted text
    txt = getattr(order, "shipping_address_text", "") or ""
    if txt.strip():
        return txt.strip()

    # 2) JSON dict from normalized schema
    data = getattr(order, "shipping_address", None)
    if isinstance(data, dict) and data:
        # show in a sensible order across supported countries
        ordered_keys = [
            "address_line1", "address_line2",
            "barangay", "city", "province", "state", "county",
            "area", "area_other", "emirate",
            "postal_code", "zip_code",
        ]
        parts = [str(data.get(k, "")).strip() for k in ordered_keys if str(data.get(k, "")).strip()]
        country = (getattr(order, "country", "") or "").strip().upper()
        if country:
            parts.append(country)
        if parts:
            return ", ".join(parts)

    # 3) legacy flat columns
    legacy = [
        getattr(order, "address_line1", "") or "",
        getattr(order, "city", "") or "",
        getattr(order, "province", "") or "",
        getattr(order, "zip_code", "") or "",
    ]
    legacy = [p for p in legacy if p]
    return ", ".join(legacy)


def _filtered_orders_queryset(request):
    status = (request.GET.get("status") or "all").strip().lower()
    q      = (request.GET.get("q") or "").strip()
    sort   = (request.GET.get("sort") or "created_desc").strip().lower()

    # Product filters
    prod_text = (request.GET.get("product") or "").strip()
    prod_id   = (request.GET.get("product_id") or "").strip()

    # Dates (inclusive)
    date_from = parse_date(request.GET.get("from") or "")
    date_to   = parse_date(request.GET.get("to") or "")

    qs = Order.objects.all()

    # Date range
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # Status
    valid_statuses = {k for k, _ in getattr(Order, "STATUS_CHOICES", ())}
    if status != "all" and status in valid_statuses:
        qs = qs.filter(status=status)

    # Global search
    if q:
        qs = qs.filter(
            Q(order_number__icontains=q) |
            Q(full_name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q)
        )

    # Product filtering (safe)
    if prod_text or prod_id:
        items_rel = _order_items_rel_name()              # 'items' in your models
        item_model = Order._meta.get_field(items_rel).related_model
        has_prod_fk = _item_has_product_fk(item_model)

        prod_q = Q()

        if prod_id and has_prod_fk:
            try:
                prod_q |= Q(**{f"{items_rel}__product__id": int(prod_id)})
            except ValueError:
                pass

        if prod_text:
            prod_q |= Q(**{f"{items_rel}__name__icontains": prod_text})
            if has_prod_fk:
                prod_q |= Q(**{f"{items_rel}__product__name__icontains": prod_text})
                prod_model = item_model._meta.get_field('product').related_model
                if _model_has_field(prod_model, 'sku'):
                    prod_q |= Q(**{f"{items_rel}__product__sku__icontains": prod_text})

        try:
            qs = qs.filter(prod_q).distinct()
        except Exception:
            # super safe fallback
            if prod_text:
                qs = qs.filter(**{f"{items_rel}__name__icontains": prod_text}).distinct()

    # Sorting
    status_order_list = [k for k, _ in getattr(Order, "STATUS_CHOICES", ())]
    status_case = Case(
        *[When(status=val, then=pos) for pos, val in enumerate(status_order_list, start=1)],
        default=len(status_order_list) + 1,
        output_field=IntegerField(),
    )

    if sort == "status_asc":
        qs = qs.order_by(status_case, "-created_at")
    elif sort == "status_desc":
        qs = qs.order_by(-status_case, "-created_at")
    elif sort == "total_asc":
        qs = qs.order_by("grand_total", "-created_at")
    elif sort == "total_desc":
        qs = qs.order_by("-grand_total", "-created_at")
    elif sort == "created_asc":
        qs = qs.order_by("created_at")
    else:
        qs = qs.order_by("-created_at")

    ctxbits = {
        "status": status,
        "q": q,
        "sort": sort,
        "product": prod_text,
        "product_id": prod_id,
        "date_from": date_from,
        "date_to": date_to,
    }
    return qs, ctxbits


# -------------------------------------------------------------------
# Auth gate (Django auth OR shared password)
# -------------------------------------------------------------------
ORDER_STATUSES = [key for key, _ in getattr(Order, "STATUS_CHOICES", ())]
DASH_AUTH_SESSION_KEY = "dashboard_authed"

def dashboard_required(viewfunc):
    @wraps(viewfunc)
    def _wrapped(request, *args, **kwargs):
        if request.user.is_authenticated or request.session.get(DASH_AUTH_SESSION_KEY):
            return viewfunc(request, *args, **kwargs)
        next_url = request.get_full_path()
        return redirect(f"/dashboard/login/?next={next_url}")
    return _wrapped

def dashboard_login(request):
    """
    Accept either:
      1) Django auth (username + password)
      2) Shared gate password via settings.DASHBOARD_PASSWORD (put it in 'password'; leave username blank)
    """
    pwd_setting = getattr(settings, "DASHBOARD_PASSWORD", "changeme")
    context = {"next": request.GET.get("next", "/dashboard/")}

    if request.method == "POST":
        next_url = request.POST.get("next") or "/dashboard/"
        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()

        if username:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                dj_login(request, user)
                messages.success(request, f"Welcome back, {user.get_username()}.")
                return redirect(next_url)
            messages.error(request, "Invalid username or password.")
            return render(request, "dashboard/login.html", context, status=200)

        if password == pwd_setting:
            request.session[DASH_AUTH_SESSION_KEY] = True
            messages.success(request, "Welcome back.")
            return redirect(next_url)

        messages.error(request, "Incorrect password.")
        return render(request, "dashboard/login.html", context, status=200)

    return render(request, "dashboard/login.html", context)

def dashboard_logout(request):
    request.session.pop(DASH_AUTH_SESSION_KEY, None)
    dj_logout(request)
    messages.info(request, "You have been logged out.")
    return redirect("/dashboard/login/")

# -------------------------------------------------------------------
# Dashboard pages
# -------------------------------------------------------------------
@dashboard_required
def dashboard_home(request):
    kpi = {
        "pending": Order.objects.filter(status="pending").count(),
        "today": Order.objects.filter(created_at__date=timezone.localdate()).count(),
        "sales_30": Order.objects.filter(
            created_at__gte=timezone.now() - timezone.timedelta(days=30)
        ).aggregate(total=Sum("grand_total"))["total"] or 0,
    }
    recent = Order.objects.order_by("-created_at")[:10]
    return render(request, "dashboard/home.html", {
        "kpi": kpi,
        "recent": recent,
        "order_statuses": ORDER_STATUSES,
    })

# myApp/views_dashboard.py
from django.db.models import Sum, Q, Case, When, IntegerField  # make sure these are imported

@dashboard_required
def order_list(request):
    qs, ctxbits = _filtered_orders_queryset(request)

    # supply products for the dropdown
    products = list(
        Product.objects.filter(is_active=True)
        .order_by('name')
        .values('id', 'name')[:1000]   # cap for safety; adjust if needed
    )

    ctx = {
        "orders": qs[:200],
        "order_statuses": [k for k, _ in getattr(Order, "STATUS_CHOICES", ())],
        "products": products,
        **ctxbits,
    }

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "dashboard/_orders_table.html", ctx)
    # Use CMS template if coming from CMS dashboard
    if '/cms/' in request.path or request.path == '/dashboard/cms/orders/':
        return render(request, "dashboard/cms_orders.html", ctx)
    return render(request, "dashboard/orders.html", ctx)

# views_dashboard.py
@dashboard_required
def order_detail(request, order_number):
    order = get_object_or_404(Order, order_number=order_number)

    items_mgr = getattr(order, "items", None)
    if hasattr(items_mgr, "all"):
        items_qs = items_mgr.all()
    else:
        items_qs = order.orderitem_set.all()

    # Best-effort select_related on product
    try:
        model = items_qs.model
        if any(getattr(f, "name", "") == "product" and getattr(f, "is_relation", False)
               for f in model._meta.get_fields()):
            items_qs = items_qs.select_related("product")
    except Exception:
        pass

    # ✅ supply shipping_text expected by the template
    shipping_text = _shipping_text_for(order)

    template = "dashboard/order_detail.html"
    # Use CMS template if coming from CMS dashboard
    if '/cms/' in request.path or request.META.get('HTTP_REFERER', '').startswith(request.build_absolute_uri('/dashboard/cms/')):
        template = "dashboard/cms_order_detail.html"

    return render(
        request,
        template,
        {
            "order": order,
            "items": items_qs,
            "order_statuses": tuple(getattr(Order, "STATUS_CHOICES", ())),
            "shipping_text": shipping_text,
        },
    )



# views_dashboard.py
@dashboard_required
def order_summary_json(request, order_number):
    order = get_object_or_404(Order, order_number=order_number)

    items_mgr = getattr(order, "items", None)
    items_qs = items_mgr.all() if hasattr(items_mgr, "all") else order.orderitem_set.all()

    placed_dt = timezone.localtime(order.created_at)
    updated_dt = timezone.localtime(getattr(order, "updated_at", order.created_at))

    try:
        status_label = order.get_status_display()
    except Exception:
        status_label = (order.status or "updated").title()

    ship = order.shipping_cost or Decimal("0")
    ship_display = _money(ship, request) if ship != 0 else "Free"

    disc = order.discount_total or Decimal("0")
    disc_display = f"-{_money(disc, request)}" if disc != 0 else ""

    return JsonResponse({
        "ok": True,
        "order": {
            "order_number": order.order_number,
            "full_name": order.full_name or "",
            "email": order.email or "",
            "phone": order.phone or "",
            # legacy fields kept for backwards compatibility
            "address_line1": getattr(order, "address_line1", "") or "",
            "city": getattr(order, "city", "") or "",
            "province": getattr(order, "province", "") or "",
            "zip_code": getattr(order, "zip_code", "") or "",
            # ✅ add normalized view
            "country": (getattr(order, "country", "") or "").upper(),
            "shipping_text": _shipping_text_for(order),
            "status": order.status,
            "status_display": status_label,
            "tracking_number": getattr(order, "tracking_number", "") or "",
            "notes": order.notes or "",
            "placed_display": placed_dt.strftime("%b %d, %Y %H:%M"),
            "updated_display": updated_dt.strftime("%b %d, %Y %H:%M"),
        },
        "totals": {
            "subtotal": str(order.subtotal or Decimal("0")),
            "subtotal_display": _money(order.subtotal or 0, request),
            "shipping": str(ship),
            "shipping_display": ship_display,
            "discount": str(disc),
            "discount_display": disc_display,
            "grand_total": str(order.grand_total or Decimal("0")),
            "grand_total_display": _money(order.grand_total or 0, request),
        },
        "items": [{
            "name": it.name,
            "qty": it.quantity,
            "unit_display": _money(getattr(it, "unit_price", 0) or 0, request),
            "total_display": _money(getattr(it, "line_total", 0) or 0, request),
        } for it in items_qs],
    })


# -------------------------------------------------------------------
# Update / Delete
# -------------------------------------------------------------------
@dashboard_required
@require_POST
def order_update(request, order_number):
    """
    Update status / tracking / notes (and cancellation reasons).
    - If moving to 'canceled', a cancel_reason is REQUIRED.
    - If cancel_reason is provided even without a status change, we'll persist it.
    - If your OrderItem has `cancel_reason`, you can pass item-level reasons via
      inputs named `item_reason_<item.id>`; they'll be saved when present.
    """
    order = get_object_or_404(Order, order_number=order_number)

    new_status    = request.POST.get("status", order.status)
    new_tracking  = (request.POST.get("tracking") or "").strip()
    new_notes     = (request.POST.get("notes") or "").strip()
    cancel_reason = (request.POST.get("cancel_reason") or "").strip()   # NEW
    next_url      = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/dashboard/"

    # Validate status
    valid_statuses = {k for k, _ in getattr(Order, "STATUS_CHOICES", ())}
    if new_status not in valid_statuses:
        messages.error(request, "Invalid status.")
        return redirect(next_url)

    # If transitioning into 'canceled', require a reason
    transitioning_to_canceled = (order.status != "canceled" and new_status == "canceled")
    if transitioning_to_canceled and not cancel_reason:
        # As a softer fallback, accept any provided item-level reasons as satisfying the requirement
        # (in case you capture reasons per product). If none found, block the update.
        has_item_reason = False
        try:
            items_qs = order.items.all() if hasattr(order, "items") else order.orderitem_set.all()
            for it in items_qs:
                val = (request.POST.get(f"item_reason_{it.id}") or "").strip()
                if val:
                    has_item_reason = True
                    break
        except Exception:
            pass

        if not has_item_reason:
            messages.error(request, "Please provide a reason for cancelling this order.")
            return redirect(next_url)

    # Detect changes
    status_changed   = (order.status != new_status)
    tracking_changed = (getattr(order, "tracking_number", "") or "") != new_tracking
    notes_changed    = (order.notes or "") != new_notes

    # Apply core fields
    order.status = new_status
    if hasattr(order, "tracking_number"):
        order.tracking_number = new_tracking
    order.notes = new_notes

    # Persist order-level cancel_reason if:
    # - we are transitioning to canceled (required above), OR
    # - a value was provided (even if status didn't change)
    if cancel_reason or transitioning_to_canceled:
        # keep the latest non-empty reason; if empty and not transitioning, don't overwrite existing
        if cancel_reason:
            # Make sure your Order model has `cancel_reason` (TextField blank=True)
            if hasattr(order, "cancel_reason"):
                order.cancel_reason = cancel_reason

    # Save now if any changed (status/tracking/notes or cancel_reason change)
    dirty = status_changed or tracking_changed or notes_changed
    # Consider cancel_reason as a change if provided and field exists
    if hasattr(order, "cancel_reason") and cancel_reason and (order.cancel_reason != cancel_reason):
        dirty = True

    if dirty:
        order.save()
        messages.success(request, f"Order {order.order_number} updated.")
    else:
        messages.info(request, "No changes made.")

    # Optional: save per-item cancellation reasons when provided
    # (safe even if your OrderItem doesn't have cancel_reason)
    saved_item_notes = False
    try:
        items_qs = order.items.all() if hasattr(order, "items") else order.orderitem_set.all()
        for it in items_qs:
            key = f"item_reason_{it.id}"
            val = (request.POST.get(key) or "").strip()
            if val and hasattr(it, "cancel_reason"):
                if (it.cancel_reason or "") != val:
                    it.cancel_reason = val
                    it.save(update_fields=["cancel_reason"])
                    saved_item_notes = True
    except Exception:
        # Don't block the flow if item-level saving fails
        pass
    if saved_item_notes:
        messages.info(request, "Saved item-level cancellation notes.")

    # Notify customer if enabled and if status or tracking changed
    notify_flag = request.POST.get("notify", "on").lower() not in ("", "0", "false", "off")
    if order.email and notify_flag and (status_changed or tracking_changed):
        _email_order_status_update(
            request, order,
            status_changed=status_changed,
            tracking_changed=tracking_changed
        )

    return redirect(next_url)


def _email_order_status_update(request, order, status_changed=True, tracking_changed=False):
    """
    Compose + send status/tracking update via your Resend-backed shim.
    """
    if not order.email:
        return

    try:
        status_label = order.get_status_display()
    except Exception:
        status_label = (order.status or "Updated").title()

    subject = f"Update: Your SHARP Order {order.order_number} is now {status_label}"

    context = {
        "order": order,
        "request": request,
        "status_changed": status_changed,
        "tracking_changed": tracking_changed,
    }

    text_body = render_to_string("emails/order_status_update.txt", context)
    html_body = render_to_string("emails/order_status_update.html", context)

    ok = _safe_send_mail(
        subject=subject,
        text_body=text_body,
        from_email=None,  # your shim uses settings.RESEND['FROM'] / DEFAULT_FROM_EMAIL
        to_list=[order.email],
        html_body=html_body,
        extra_headers={
            "Reply-To": (
                getattr(settings, "CONTACT_TO", "")
                or getattr(settings, "DEFAULT_FROM_EMAIL", "")
                or ""
            )
        },
    )

    if not ok:
        messages.warning(request, "Order updated, but we couldn’t send the status email. We’ll resend shortly.")

from django.urls import reverse

@dashboard_required
@require_POST
def order_delete(request, order_number):
    """
    Delete a single order.
    - Requires a cancel_reason (so you keep context).
    - Stores the reason on the order before deletion (if `cancel_reason` field exists).
    - Always redirects to dashboard_home after delete.
    """
    reason = (request.POST.get("cancel_reason") or "").strip()
    if not reason:
        messages.error(request, "Please provide a reason for deletion.")
        # send them back to where they came from, or to the detail page if available
        return redirect(
            request.META.get("HTTP_REFERER")
            or reverse("dashboard_order_detail", args=[order_number])
        )

    order = get_object_or_404(Order, order_number=order_number)

    # Persist the reason for audit before delete (if your model has the field)
    if hasattr(order, "cancel_reason"):
        order.cancel_reason = reason
        order.save(update_fields=["cancel_reason"])

    order.delete()
    messages.info(request, f"Order {order_number} deleted.")

    # ✅ Always go to dashboard home after deletion
    return redirect(reverse("dashboard_home"))



from django.views.decorators.http import require_POST

@dashboard_required
@require_POST
def order_bulk_delete(request):
    """
    Deletes multiple orders by order_number. Expects a POST with:
      - order_numbers: CSV list of order numbers (e.g. "SH-000001,SH-000002")
      - next: optional URL to redirect back to
    """
    raw = (request.POST.get("order_numbers") or "").strip()
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/dashboard/"

    # Parse CSV -> list
    order_numbers = [s.strip() for s in raw.split(",") if s.strip()]
    if not order_numbers:
        messages.info(request, "No orders selected.")
        return redirect(next_url)

    qs = Order.objects.filter(order_number__in=order_numbers)
    count = qs.count()
    if not count:
        messages.info(request, "No matching orders found.")
        return redirect(next_url)

    qs.delete()  # cascades to items; that's fine
    messages.success(request, f"Deleted {count} order(s).")
    return redirect(next_url)



from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q

@dashboard_required
def order_export(request):
    """
    Export filtered orders as Excel (.xlsx) or PDF.
    Rows are per line-item (order can appear multiple times).
    Filters honored: status, sort, q, from, to, product, product_id
    """
    export_fmt = (request.GET.get("export") or "excel").lower()
    qs, _ = _filtered_orders_queryset(request)

    # Prefetch items (+ product if FK exists)
    items_rel = _order_items_rel_name()
    item_model = Order._meta.get_field(items_rel).related_model
    has_prod_fk = _item_has_product_fk(item_model)
    prefetch_path = f"{items_rel}__product" if has_prod_fk else items_rel
    qs = qs.prefetch_related(prefetch_path)

    prod_text = (request.GET.get("product") or "").strip()
    prod_id   = (request.GET.get("product_id") or "").strip()

    # Build flat rows (one row per item)
    rows = []
    for order in qs[:5000]:  # safety cap
        items_mgr = getattr(order, "items", None)
        items_qs = items_mgr.all() if hasattr(items_mgr, "all") else order.orderitem_set.all()

        # If product filter applied, narrow items for this order too (export should match list)
        if prod_text or prod_id:
            try:
                if prod_id and has_prod_fk:
                    items_qs = items_qs.filter(product_id=int(prod_id))
            except ValueError:
                pass
            if prod_text:
                name_q = Q(name__icontains=prod_text)
                if has_prod_fk:
                    name_q |= Q(product__name__icontains=prod_text)
                    # SKU optional
                    prod_model = item_model._meta.get_field('product').related_model
                    if _model_has_field(prod_model, 'sku'):
                        name_q |= Q(product__sku__icontains=prod_text)
                items_qs = items_qs.filter(name_q)

        for it in items_qs:
            item_name = (getattr(it, "name", None) or
                         getattr(getattr(it, "product", None), "name", None) or "")[:255]
            sku = getattr(getattr(it, "product", None), "sku", "") or ""
            qty = int(getattr(it, "quantity", 0) or 0)
            unit_display = _money(getattr(it, "unit_price", 0) or 0, request)
            line_display = _money(getattr(it, "line_total", 0) or 0, request)

            try:
                status_label = order.get_status_display()
            except Exception:
                status_label = (order.status or "").title()

            rows.append({
                "Order #": order.order_number,
                "Date": timezone.localtime(order.created_at).strftime("%Y-%m-%d %H:%M"),
                "Status": status_label,
                "Customer": order.full_name or "",
                "Email": order.email or "",
                "Phone": order.phone or "",
                "Item": item_name,
                "SKU": sku,
                "Qty": qty,
                "Unit Price": unit_display,
                "Line Total": line_display,
                "Grand Total": _money(order.grand_total or 0, request),
            })

    if export_fmt == "pdf":
        return _export_pdf(rows)

    # default: Excel
    return _export_xlsx(rows)


def _export_xlsx(rows):
    """
    Build and return an .xlsx response.
    If openpyxl isn't installed, return a 400 with a friendly message.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse(
            "Excel export is not available (openpyxl not installed). "
            "Add 'openpyxl' to requirements.",
            status=400
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ["Order #","Date","Status","Customer","Email","Phone","Item","SKU","Qty","Unit Price","Line Total","Grand Total"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # autosize
    for col_idx, h in enumerate(headers, start=1):
        max_len = max(
            [len(str(h))] +
            [len(str(ws.cell(row=i, column=col_idx).value or "")) for i in range(1, ws.max_row+1)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(50, max_len + 2)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    resp = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
    return resp


def _export_pdf(rows):
    """
    Build and return a PDF response.
    If reportlab isn't installed, return a 400 with a friendly message.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return HttpResponse(
            "PDF export is not available (reportlab not installed). "
            "Add 'reportlab' to requirements.",
            status=400
        )

    buff = BytesIO()
    doc = SimpleDocTemplate(
        buff, pagesize=landscape(A4),
        leftMargin=18, rightMargin=18, topMargin=24, bottomMargin=24
    )

    styles = getSampleStyleSheet()
    story = [Paragraph("Orders Export", styles["Title"]), Spacer(1, 8)]

    headers = ["Order #","Date","Status","Customer","Email","Phone","Item","SKU","Qty","Unit Price","Line Total","Grand Total"]
    data = [headers]
    for r in rows:
        data.append([str(r.get(h, "")) for h in headers])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#263128")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 10),
        ("GRID",       (0,0), (-1,-1), 0.25, colors.HexColor("#E1E1E1")),
        ("FONTSIZE",   (0,1), (-1,-1), 8),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.HexColor("#F7F7F7")]),
    ]))

    story.append(table)
    doc.build(story)

    pdf = buff.getvalue()
    buff.close()
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="orders.pdf"'
    return resp



from django import forms
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.contrib import messages
from django.utils import timezone
from .models import PromoCode

# views_dashboard.py (or forms.py if you split it)
from django import forms
from .models import PromoCode

class PromoForm(forms.ModelForm):
    code = forms.CharField(help_text="e.g., SHARP10 (letters/numbers only)")

    class Meta:
        model = PromoCode
        fields = [
            "code","description","type","value","min_subtotal","max_discount",
            "countries_csv","starts_at","ends_at","active","is_sitewide","usage_limit",
        ]
        widgets = {
            "starts_at": forms.DateTimeInput(attrs={"type": "datetime-local"}),
            "ends_at": forms.DateTimeInput(attrs={"type": "datetime-local"}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        base = "mt-1 w-full px-3 py-2 rounded-xl border border-[#E1E1E1] bg-white"
        for name, field in self.fields.items():
            # Don’t style checkboxes like text inputs
            if isinstance(field.widget, (forms.CheckboxInput,)):
                continue
            cls = field.widget.attrs.get("class", "")
            field.widget.attrs["class"] = f"{cls} {base}".strip()

    def clean_code(self):
        c = (self.cleaned_data["code"] or "").strip().upper()
        if not c.replace("-", "").isalnum():
            raise forms.ValidationError("Use letters/numbers and optional dashes only.")
        return c

    def clean(self):
        cleaned_data = super().clean()
        is_sitewide = cleaned_data.get("is_sitewide", False)
        active = cleaned_data.get("active", False)
        
        # If this promo is being set as sitewide and active, check for conflicts
        if is_sitewide and active:
            # Exclude current instance if editing
            instance = self.instance
            existing = PromoCode.objects.filter(
                is_sitewide=True,
                active=True
            ).exclude(pk=instance.pk if instance.pk else None)
            
            # Check if any existing sitewide promos are live (considering dates/usage)
            from django.utils import timezone
            now = timezone.now()
            for promo in existing:
                if promo.is_live():
                    raise forms.ValidationError(
                        f"Another sitewide promo ({promo.code}) is already active. "
                        "Only one sitewide promo can be active at a time. "
                        "Please deactivate the other promo first."
                    )
        
        return cleaned_data


@dashboard_required
def promo_list(request):
    q = (request.GET.get("q") or "").strip()
    promos = PromoCode.objects.all()
    if q:
        promos = promos.filter(code__icontains=q)
    # Use CMS template if coming from CMS dashboard
    if '/cms/' in request.path or request.path == '/dashboard/cms/promos/':
        return render(request, "dashboard/cms_promos.html", {"promos": promos, "q": q})
    return render(request, "dashboard/promos/list.html", {"promos": promos, "q": q})

@dashboard_required
def promo_upsert(request, pk=None):
    promo = get_object_or_404(PromoCode, pk=pk) if pk else None
    if request.method == "POST":
        form = PromoForm(request.POST, instance=promo)
        if form.is_valid():
            obj = form.save(commit=False)
            # normalize code upper
            obj.code = obj.code.upper()
            obj.save()
            messages.success(request, f"Promo {obj.code} saved.")
            # Redirect to CMS if coming from CMS
            if '/cms/' in request.path or request.META.get('HTTP_REFERER', '').startswith(request.build_absolute_uri('/dashboard/cms/')):
                return redirect("cms_promo_list")
            return redirect("dashboard_promo_list")
    else:
        form = PromoForm(instance=promo)
    # Use CMS template if coming from CMS dashboard
    if '/cms/' in request.path or request.META.get('HTTP_REFERER', '').startswith(request.build_absolute_uri('/dashboard/cms/')):
        return render(request, "dashboard/cms_promo_form.html", {"form": form, "promo": promo})
    return render(request, "dashboard/promos/form.html", {"form": form, "promo": promo})

@dashboard_required
def promo_toggle(request, pk):
    promo = get_object_or_404(PromoCode, pk=pk)
    promo.active = not promo.active
    promo.save(update_fields=["active","updated_at"])
    messages.info(request, f"{promo.code} is now {'active' if promo.active else 'inactive'}.")
    # Redirect to CMS if coming from CMS
    if '/cms/' in request.path or request.META.get('HTTP_REFERER', '').startswith(request.build_absolute_uri('/dashboard/cms/')):
        return redirect("cms_promo_list")
    return redirect("dashboard_promo_list")

@dashboard_required
def promo_delete(request, pk):
    promo = get_object_or_404(PromoCode, pk=pk)
    code = promo.code
    promo.delete()
    messages.warning(request, f"Promo {code} deleted.")
    # Redirect to CMS if coming from CMS
    if '/cms/' in request.path or request.META.get('HTTP_REFERER', '').startswith(request.build_absolute_uri('/dashboard/cms/')):
        return redirect("cms_promo_list")
    return redirect("dashboard_promo_list")


from django.shortcuts import get_object_or_404, render
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.safestring import mark_safe
import json

# at top
from myApp.views import _physical_items_for_order, WASSEL_CODE_LABELS

@dashboard_required
def wasselexpress_preview(request, order_number):
    order = get_object_or_404(Order, order_number=order_number)
    addr = dict(order.shipping_address or {})
    car  = dict(addr.get("_carrier") or {})
    hist = list(car.get("history") or [])
    last = hist[-1] if hist else None
    carrier = dict(addr.get("_carrier") or {}) 
    code = (car.get("last_update") or str(order.status) or "").strip()
    label = WASSEL_CODE_LABELS.get(int(code)) if code.isdigit() else code or "—"

    picklist = list(_physical_items_for_order(order).items())  # [(name, qty), ...]

    ctx = {
        "order": order,
        "carrier": {
            "awb": car.get("awb"),
            "status_code": code or "0",
            "status_label": label or "Created",
            "companyStoreID": getattr(settings, "WASSEL", {}).get("COMPANY_STORE_ID", 13),
            "recipientCity": addr.get("city") or addr.get("area") or addr.get("emirate") or "",
            "recipientArea": addr.get("area") or addr.get("barangay") or "",
            "addressDescription": ", ".join([str(addr.get(k) or "") for k in
                ("address_line1","address_line2","area","barangay","city","province","state","postal_code") if addr.get(k)]),
            "recipientName": order.full_name,
            "recipientEmail": order.email,
            "recipientPhoneNumber": order.phone,
            "codAmount": ("0" if (order.payment_method or "").lower() != "cod" else str(order.grand_total)),
            "referenceID": order.order_number,
            "history": hist,
        },
        "picklist": picklist,
        "raw_response": json.dumps(car.get("raw") or {}, indent=2, ensure_ascii=False),
        "payload": carrier.get("last_payload"),
    }
    return render(request, "dashboard/wsl_preview.html", ctx)




# myApp/views_dashboard.py
# --- add near other imports ---
from django import forms
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.core.paginator import Paginator

from .models import Product, ProductComponent

# -----------------------------
# Forms (styled like your PromoForm)
# -----------------------------
class ProductForm(forms.ModelForm):
    class Meta:
        model = Product
        fields = [
            "name","sku","short_description","description",
            "price","image_url","gallery_csv",
            "is_active","is_bundle","free_delivery",
        ]
        widgets = {
            "description": forms.Textarea(attrs={"rows": 4}),
        }

    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        base = "mt-1 w-full px-3 py-2 rounded-xl border border-[#E1E1E1] bg-white"
        for name, field in self.fields.items():
            if isinstance(field.widget, (forms.CheckboxInput,)):
                continue
            cls = field.widget.attrs.get("class","")
            field.widget.attrs["class"] = f"{cls} {base}".strip()

    def clean_price(self):
        val = self.cleaned_data.get("price")
        if val is None or val < 0:
            raise forms.ValidationError("Price must be zero or higher.")
        return val


class BundleComponentsField(forms.CharField):
    """
    Accepts rows like:
      SKU-123 x2
      Conditioner 500ml x1
    or CSV: "SKU-123,2" per line. Name or SKU is matched to Product.
    """
    def clean(self, value):
        txt = (value or "").strip()
        rows = []
        for line in txt.splitlines():
            line = line.strip()
            if not line:
                continue
            # try "name xQTY"
            qty = 1
            if " x" in line.lower():
                base, q = line.rsplit(" x", 1)
                try:
                    qty = max(1, int(q.strip()))
                    key = base.strip()
                except Exception:
                    key = line
            else:
                # try CSV "name,qty"
                if "," in line:
                    base, q = line.split(",",1)
                    key = base.strip()
                    try:
                        qty = max(1, int(q.strip()))
                    except Exception:
                        qty = 1
                else:
                    key = line
            rows.append((key, qty))
        return rows


class BundleForm(forms.Form):
    components = BundleComponentsField(
        required=False,
        help_text="One per line, e.g. 'SKU-123 x2' or 'Conditioner 500ml x1'."
    )


# -----------------------------
# Product list + search
# -----------------------------
@dashboard_required
def product_list(request):
    q = (request.GET.get("q") or "").strip()
    active = (request.GET.get("active") or "all").lower()
    sort = (request.GET.get("sort") or "name").lower()

    qs = Product.objects.all()
    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(sku__icontains=q) | Q(slug__icontains=q)
        )
    if active in ("1","true","yes","active"):
        qs = qs.filter(is_active=True)
    elif active in ("0","false","no","inactive"):
        qs = qs.filter(is_active=False)

    if sort == "price_desc":
        qs = qs.order_by("-price","name")
    elif sort == "price_asc":
        qs = qs.order_by("price","name")
    elif sort == "created_desc":
        qs = qs.order_by("-created_at")
    elif sort == "created_asc":
        qs = qs.order_by("created_at")
    else:
        qs = qs.order_by("name")

    page = Paginator(qs, 50).get_page(request.GET.get("page"))

    ctx = {"page": page, "q": q, "active": active, "sort": sort}
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "dashboard/products/_table.html", ctx)
    # Use CMS template if coming from CMS dashboard
    if '/cms/' in request.path or request.path == '/dashboard/cms/products/':
        return render(request, "dashboard/cms_products.html", ctx)
    return render(request, "dashboard/products/list.html", ctx)


# -----------------------------
# Upsert (modal form)
# -----------------------------
# views_dashboard.py (inside product_upsert)
from django.http import JsonResponse

@dashboard_required
def product_upsert(request, pk=None):
    product = get_object_or_404(Product, pk=pk) if pk else None

    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        bundle_form = BundleForm(request.POST)  # or however you build it
        if form.is_valid() and (not form.cleaned_data.get("is_bundle") or bundle_form.is_valid()):
            obj = form.save()
            # TODO: persist bundle_form if applicable
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"ok": True})
            messages.success(request, f"Product '{obj.name}' saved.")
            # Redirect to CMS if coming from CMS
            if '/cms/' in request.path or request.META.get('HTTP_REFERER', '').startswith(request.build_absolute_uri('/dashboard/cms/')):
                return redirect("cms_product_list")
            return redirect("dashboard_product_list")
        # invalid -> fall through to render with errors
    else:
        form = ProductForm(instance=product)
        bundle_form = BundleForm(initial={})  # or None, up to you

    ctx = {"form": form, "bundle_form": bundle_form, "product": product}

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "dashboard/products/form_modal.html", ctx)

    return render(request, "dashboard/products/form.html", ctx)


# -----------------------------
# Quick actions (AJAX-friendly)
# -----------------------------
@dashboard_required
@require_POST
def product_toggle_active(request, pk):
    prod = get_object_or_404(Product, pk=pk)
    prod.is_active = not prod.is_active
    prod.save(update_fields=["is_active"])
    return JsonResponse({"ok": True, "is_active": prod.is_active})


@dashboard_required
@require_POST
def product_update_price(request, pk):
    prod = get_object_or_404(Product, pk=pk)
    try:
        new_price = Decimal((request.POST.get("price") or "").strip())
        if new_price < 0:
            raise ValueError
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid price."}, status=400)
    prod.price = new_price
    prod.save(update_fields=["price"])
    return JsonResponse({"ok": True, "price": f"{prod.price:.2f}"})


from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render, redirect

def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.delete()
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": True})
        # Redirect to CMS if coming from CMS
        if '/cms/' in request.path or request.META.get('HTTP_REFERER', '').startswith(request.build_absolute_uri('/dashboard/cms/')):
            return redirect("cms_product_list")
        return redirect("dashboard_product_list")

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "dashboard/products/_delete_confirm.html", {"product": product})
    return redirect("dashboard_product_list")


# -----------------------------
# Bulk actions
# -----------------------------
@dashboard_required
@require_POST
def product_bulk_action(request):
    """
    POST:
      ids: CSV of product IDs
      action: activate|deactivate|price_percent
      percent: optional (for price_percent, e.g. 10 or -5)
    """
    raw = (request.POST.get("ids") or "").strip()
    ids = [int(x) for x in raw.split(",") if x.strip().isdigit()]
    qs = Product.objects.filter(id__in=ids)
    action = (request.POST.get("action") or "").strip()

    if not ids:
        return JsonResponse({"ok": False, "error": "No items selected."}, status=400)

    count = 0
    if action == "activate":
        count = qs.update(is_active=True)
    elif action == "deactivate":
        count = qs.update(is_active=False)
    elif action == "price_percent":
        try:
            pct = Decimal((request.POST.get("percent") or "0").strip())
        except Exception:
            return JsonResponse({"ok": False, "error": "Invalid percent."}, status=400)
        with transaction.atomic():
            for p in qs.select_for_update():
                p.price = (p.price * (Decimal("1.00") + pct/Decimal("100"))).quantize(Decimal("0.01"))
                p.save(update_fields=["price"])
                count += 1
    else:
        return JsonResponse({"ok": False, "error": "Unknown action."}, status=400)

    return JsonResponse({"ok": True, "count": count})


# ============================================================================
# CMS DASHBOARD VIEWS
# ============================================================================

from .models import (
    MediaAsset, SEO, Navigation, Hero, About, Stat, Service,
    Portfolio, PortfolioProject, Testimonial, FAQ, FAQItem,
    Contact, ContactInfo, ContactFormField, SocialLink, Footer
)
from .utils.cloudinary_utils import upload_to_cloudinary
from django.views.decorators.csrf import csrf_exempt
import json


@dashboard_required
def cms_dashboard_home(request):
    """CMS Dashboard Homepage with KPIs"""
    # KPIs
    kpi = {
        "pending": Order.objects.filter(status="0").count(),
        "today": Order.objects.filter(created_at__date=timezone.localdate()).count(),
        "sales_30": Order.objects.filter(
            created_at__gte=timezone.now() - timezone.timedelta(days=30)
        ).aggregate(total=Sum("grand_total"))["total"] or 0,
    }
    
    recent = Order.objects.order_by("-created_at")[:10]
    
    # CMS Stats
    try:
        media_count = MediaAsset.objects.count()
    except:
        media_count = 0
    
    try:
        navigation_count = Navigation.objects.filter(is_active=True).count()
    except:
        navigation_count = 0
    
    try:
        services_count = Service.objects.filter(is_active=True).count()
    except:
        services_count = 0
    
    try:
        testimonials_count = Testimonial.objects.filter(is_active=True).count()
    except:
        testimonials_count = 0
    
    try:
        product_count = Product.objects.filter(is_active=True).count()
    except:
        product_count = 0
    
    context = {
        'kpi': kpi,
        'recent': recent,
        'order_statuses': getattr(Order, "STATUS_CHOICES", []),
        'media_count': media_count,
        'navigation_count': navigation_count,
        'services_count': services_count,
        'testimonials_count': testimonials_count,
        'product_count': product_count,
    }
    return render(request, 'dashboard/cms_index.html', context)


# ============================================================================
# Image Upload & Gallery
# ============================================================================

@dashboard_required
@csrf_exempt
def upload_image(request):
    """Upload image to Cloudinary"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    
    if 'image' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No image file'}, status=400)
    
    image_file = request.FILES['image']
    folder = request.POST.get('folder', 'uploads')
    
    result = upload_to_cloudinary(image_file, folder=folder)
    
    if result['success']:
        # Save to MediaAsset
        asset = MediaAsset.objects.create(
            title=request.POST.get('title', ''),
            original_url=result['original_url'],
            web_url=result['web_url'],
            thumbnail_url=result['thumbnail_url'],
            folder=folder,
            public_id=result['public_id'],
            width=result.get('width'),
            height=result.get('height'),
            file_size=result.get('bytes'),
        )
        return JsonResponse({
            'success': True,
            'url': result['original_url'],
            'web_url': result['web_url'],
            'thumbnail_url': result['thumbnail_url'],
            'asset_id': asset.id,
        })
    else:
        return JsonResponse({'success': False, 'error': result.get('error', 'Upload failed')}, status=500)


@dashboard_required
def gallery(request):
    """Image gallery"""
    folder = request.GET.get('folder', '')
    assets = MediaAsset.objects.all()
    if folder:
        assets = assets.filter(folder=folder)
    assets = assets.order_by('-created_at')
    
    folders = MediaAsset.objects.values_list('folder', flat=True).distinct()
    
    context = {
        'assets': assets,
        'folders': folders,
        'current_folder': folder,
    }
    return render(request, 'dashboard/gallery.html', context)


# ============================================================================
# SEO
# ============================================================================

@dashboard_required
def seo_edit(request):
    """Edit SEO settings"""
    seo, created = SEO.objects.get_or_create(page='home')
    
    if request.method == 'POST':
        seo.title = request.POST.get('title', '')
        seo.description = request.POST.get('description', '')
        seo.keywords = request.POST.get('keywords', '')
        seo.og_image_url = request.POST.get('og_image_url', '')
        seo.save()
        messages.success(request, 'SEO settings updated successfully.')
        return redirect('dashboard:seo_edit')
    
    context = {'seo': seo}
    return render(request, 'dashboard/seo_edit.html', context)


# ============================================================================
# Navigation
# ============================================================================

@dashboard_required
def navigation_edit(request):
    """Edit navigation menu"""
    items = Navigation.objects.all().order_by('sort_order')
    
    if request.method == 'POST':
        # Handle delete
        if 'delete_id' in request.POST:
            item = get_object_or_404(Navigation, id=request.POST['delete_id'])
            item.delete()
            messages.success(request, 'Navigation item deleted.')
            return redirect('dashboard:navigation_edit')
        
        # Handle add/edit
        item_id = request.POST.get('item_id')
        if item_id:
            item = get_object_or_404(Navigation, id=item_id)
        else:
            item = Navigation()
        
        item.label = request.POST.get('label', '')
        item.url = request.POST.get('url', '')
        item.sort_order = int(request.POST.get('sort_order', 0))
        item.is_active = request.POST.get('is_active') == 'on'
        item.is_external = request.POST.get('is_external') == 'on'
        item.save()
        
        messages.success(request, 'Navigation item saved successfully.')
        return redirect('dashboard:navigation_edit')
    
    context = {'items': items}
    return render(request, 'dashboard/navigation_edit.html', context)


# ============================================================================
# Hero
# ============================================================================

@dashboard_required
def hero_edit(request):
    """Edit hero section"""
    hero, created = Hero.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        hero.title = request.POST.get('title', '')
        hero.subtitle = request.POST.get('subtitle', '')
        hero.cta_text = request.POST.get('cta_text', '')
        hero.cta_url = request.POST.get('cta_url', '')
        hero.background_image_url = request.POST.get('background_image_url', '')
        hero.is_active = request.POST.get('is_active') == 'on'
        hero.save()
        messages.success(request, 'Hero section updated successfully.')
        return redirect('dashboard:hero_edit')
    
    context = {'hero': hero}
    return render(request, 'dashboard/hero_edit.html', context)


# ============================================================================
# About
# ============================================================================

@dashboard_required
def about_edit(request):
    """Edit about section"""
    about, created = About.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        about.title = request.POST.get('title', '')
        about.content = request.POST.get('content', '')
        about.image_url = request.POST.get('image_url', '')
        about.is_active = request.POST.get('is_active') == 'on'
        about.save()
        messages.success(request, 'About section updated successfully.')
        return redirect('dashboard:about_edit')
    
    context = {'about': about}
    return render(request, 'dashboard/about_edit.html', context)


# ============================================================================
# Stats
# ============================================================================

@dashboard_required
def stats_list(request):
    """List all stats"""
    stats = Stat.objects.all().order_by('sort_order')
    context = {'stats': stats}
    return render(request, 'dashboard/stats_list.html', context)


@dashboard_required
def stat_edit(request, stat_id=None):
    """Edit a stat"""
    if stat_id:
        stat = get_object_or_404(Stat, id=stat_id)
    else:
        stat = Stat()
    
    if request.method == 'POST':
        stat.label = request.POST.get('label', '')
        stat.value = request.POST.get('value', '')
        stat.icon = request.POST.get('icon', '')
        stat.sort_order = int(request.POST.get('sort_order', 0))
        stat.is_active = request.POST.get('is_active') == 'on'
        stat.save()
        messages.success(request, 'Stat saved successfully.')
        return redirect('dashboard:stats_list')
    
    context = {'stat': stat}
    return render(request, 'dashboard/stat_edit.html', context)


@dashboard_required
@require_POST
def stat_delete(request, stat_id):
    """Delete a stat"""
    stat = get_object_or_404(Stat, id=stat_id)
    stat.delete()
    messages.success(request, 'Stat deleted successfully.')
    return redirect('dashboard:stats_list')


# ============================================================================
# Services
# ============================================================================

@dashboard_required
def services_list(request):
    """List all services"""
    services = Service.objects.all().order_by('sort_order')
    context = {'services': services}
    return render(request, 'dashboard/services_list.html', context)


@dashboard_required
def service_edit(request, service_id=None):
    """Edit a service"""
    if service_id:
        service = get_object_or_404(Service, id=service_id)
    else:
        service = Service()
    
    if request.method == 'POST':
        service.title = request.POST.get('title', '')
        service.description = request.POST.get('description', '')
        service.icon = request.POST.get('icon', '')
        service.image_url = request.POST.get('image_url', '')
        service.sort_order = int(request.POST.get('sort_order', 0))
        service.is_active = request.POST.get('is_active') == 'on'
        service.save()
        messages.success(request, 'Service saved successfully.')
        return redirect('dashboard:services_list')
    
    context = {'service': service}
    return render(request, 'dashboard/service_edit.html', context)


@dashboard_required
@require_POST
def service_delete(request, service_id):
    """Delete a service"""
    service = get_object_or_404(Service, id=service_id)
    service.delete()
    messages.success(request, 'Service deleted successfully.')
    return redirect('dashboard:services_list')


# ============================================================================
# Portfolio
# ============================================================================

@dashboard_required
def portfolio_edit(request):
    """Edit portfolio section header"""
    portfolio, created = Portfolio.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        portfolio.title = request.POST.get('title', '')
        portfolio.subtitle = request.POST.get('subtitle', '')
        portfolio.is_active = request.POST.get('is_active') == 'on'
        portfolio.save()
        messages.success(request, 'Portfolio section updated successfully.')
        return redirect('dashboard:portfolio_edit')
    
    context = {'portfolio': portfolio}
    return render(request, 'dashboard/portfolio_edit.html', context)


@dashboard_required
def portfolio_projects_list(request):
    """List all portfolio projects"""
    projects = PortfolioProject.objects.all().order_by('sort_order')
    context = {'projects': projects}
    return render(request, 'dashboard/portfolio_projects_list.html', context)


@dashboard_required
def portfolio_project_edit(request, project_id=None):
    """Edit a portfolio project"""
    if project_id:
        project = get_object_or_404(PortfolioProject, id=project_id)
    else:
        project = PortfolioProject()
    
    if request.method == 'POST':
        project.title = request.POST.get('title', '')
        project.description = request.POST.get('description', '')
        project.image_url = request.POST.get('image_url', '')
        project.project_url = request.POST.get('project_url', '')
        project.sort_order = int(request.POST.get('sort_order', 0))
        project.is_active = request.POST.get('is_active') == 'on'
        project.save()
        messages.success(request, 'Portfolio project saved successfully.')
        return redirect('dashboard:portfolio_projects_list')
    
    context = {'project': project}
    return render(request, 'dashboard/portfolio_project_edit.html', context)


@dashboard_required
@require_POST
def portfolio_project_delete(request, project_id):
    """Delete a portfolio project"""
    project = get_object_or_404(PortfolioProject, id=project_id)
    project.delete()
    messages.success(request, 'Portfolio project deleted successfully.')
    return redirect('dashboard:portfolio_projects_list')


# ============================================================================
# Testimonials
# ============================================================================

@dashboard_required
def testimonials_list(request):
    """List all testimonials"""
    testimonials = Testimonial.objects.all().order_by('sort_order')
    context = {'testimonials': testimonials}
    return render(request, 'dashboard/testimonials_list.html', context)


@dashboard_required
def testimonial_edit(request, testimonial_id=None):
    """Edit a testimonial"""
    if testimonial_id:
        testimonial = get_object_or_404(Testimonial, id=testimonial_id)
    else:
        testimonial = Testimonial()
    
    if request.method == 'POST':
        testimonial.name = request.POST.get('name', '')
        testimonial.role = request.POST.get('role', '')
        testimonial.company = request.POST.get('company', '')
        testimonial.content = request.POST.get('content', '')
        testimonial.image_url = request.POST.get('image_url', '')
        testimonial.rating = int(request.POST.get('rating', 5))
        testimonial.sort_order = int(request.POST.get('sort_order', 0))
        testimonial.is_active = request.POST.get('is_active') == 'on'
        testimonial.save()
        messages.success(request, 'Testimonial saved successfully.')
        return redirect('dashboard:testimonials_list')
    
    context = {'testimonial': testimonial}
    return render(request, 'dashboard/testimonial_edit.html', context)


@dashboard_required
@require_POST
def testimonial_delete(request, testimonial_id):
    """Delete a testimonial"""
    testimonial = get_object_or_404(Testimonial, id=testimonial_id)
    testimonial.delete()
    messages.success(request, 'Testimonial deleted successfully.')
    return redirect('dashboard:testimonials_list')


# ============================================================================
# FAQs
# ============================================================================

@dashboard_required
def faq_section_edit(request):
    """Edit FAQ section header"""
    faq, created = FAQ.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        faq.title = request.POST.get('title', '')
        faq.subtitle = request.POST.get('subtitle', '')
        faq.is_active = request.POST.get('is_active') == 'on'
        faq.save()
        messages.success(request, 'FAQ section updated successfully.')
        return redirect('dashboard:faq_section_edit')
    
    context = {'faq': faq}
    return render(request, 'dashboard/faq_section_edit.html', context)


@dashboard_required
def faqs_list(request):
    """List all FAQ items"""
    faqs = FAQItem.objects.all().order_by('sort_order')
    context = {'faqs': faqs}
    return render(request, 'dashboard/faqs_list.html', context)


@dashboard_required
def faq_edit(request, faq_id=None):
    """Edit an FAQ item"""
    if faq_id:
        faq = get_object_or_404(FAQItem, id=faq_id)
    else:
        faq = FAQItem()
    
    if request.method == 'POST':
        faq.question = request.POST.get('question', '')
        faq.answer = request.POST.get('answer', '')
        faq.sort_order = int(request.POST.get('sort_order', 0))
        faq.is_active = request.POST.get('is_active') == 'on'
        faq.save()
        messages.success(request, 'FAQ item saved successfully.')
        return redirect('dashboard:faqs_list')
    
    context = {'faq': faq}
    return render(request, 'dashboard/faq_edit.html', context)


@dashboard_required
@require_POST
def faq_delete(request, faq_id):
    """Delete an FAQ item"""
    faq = get_object_or_404(FAQItem, id=faq_id)
    faq.delete()
    messages.success(request, 'FAQ item deleted successfully.')
    return redirect('dashboard:faqs_list')


# ============================================================================
# Contact
# ============================================================================

@dashboard_required
def contact_edit(request):
    """Edit contact section header"""
    contact, created = Contact.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        contact.title = request.POST.get('title', '')
        contact.subtitle = request.POST.get('subtitle', '')
        contact.is_active = request.POST.get('is_active') == 'on'
        contact.save()
        messages.success(request, 'Contact section updated successfully.')
        return redirect('dashboard:contact_edit')
    
    context = {'contact': contact}
    return render(request, 'dashboard/contact_edit.html', context)


@dashboard_required
def contact_info_list(request):
    """List all contact info items"""
    info_items = ContactInfo.objects.all().order_by('sort_order')
    context = {'info_items': info_items}
    return render(request, 'dashboard/contact_info_list.html', context)


@dashboard_required
def contact_info_edit(request, info_id=None):
    """Edit a contact info item"""
    if info_id:
        info = get_object_or_404(ContactInfo, id=info_id)
    else:
        info = ContactInfo()
    
    if request.method == 'POST':
        info.label = request.POST.get('label', '')
        info.value = request.POST.get('value', '')
        info.icon = request.POST.get('icon', '')
        info.sort_order = int(request.POST.get('sort_order', 0))
        info.is_active = request.POST.get('is_active') == 'on'
        info.save()
        messages.success(request, 'Contact info saved successfully.')
        return redirect('dashboard:contact_info_list')
    
    context = {'info': info}
    return render(request, 'dashboard/contact_info_edit.html', context)


@dashboard_required
@require_POST
def contact_info_delete(request, info_id):
    """Delete a contact info item"""
    info = get_object_or_404(ContactInfo, id=info_id)
    info.delete()
    messages.success(request, 'Contact info deleted successfully.')
    return redirect('dashboard:contact_info_list')


@dashboard_required
def contact_form_fields_list(request):
    """List all contact form fields"""
    fields = ContactFormField.objects.all().order_by('sort_order')
    context = {'fields': fields}
    return render(request, 'dashboard/contact_form_fields_list.html', context)


@dashboard_required
def contact_form_field_edit(request, field_id=None):
    """Edit a contact form field"""
    if field_id:
        field = get_object_or_404(ContactFormField, id=field_id)
    else:
        field = ContactFormField()
    
    if request.method == 'POST':
        field.label = request.POST.get('label', '')
        field.field_type = request.POST.get('field_type', 'text')
        field.name = request.POST.get('name', '')
        field.placeholder = request.POST.get('placeholder', '')
        field.required = request.POST.get('required') == 'on'
        field.sort_order = int(request.POST.get('sort_order', 0))
        field.is_active = request.POST.get('is_active') == 'on'
        field.save()
        messages.success(request, 'Form field saved successfully.')
        return redirect('dashboard:contact_form_fields_list')
    
    context = {'field': field}
    return render(request, 'dashboard/contact_form_field_edit.html', context)


@dashboard_required
@require_POST
def contact_form_field_delete(request, field_id):
    """Delete a contact form field"""
    field = get_object_or_404(ContactFormField, id=field_id)
    field.delete()
    messages.success(request, 'Form field deleted successfully.')
    return redirect('dashboard:contact_form_fields_list')


# ============================================================================
# Social Links
# ============================================================================

@dashboard_required
def social_links_list(request):
    """List all social links"""
    links = SocialLink.objects.all().order_by('sort_order')
    context = {'links': links}
    return render(request, 'dashboard/social_links_list.html', context)


@dashboard_required
def social_link_edit(request, link_id=None):
    """Edit a social link"""
    if link_id:
        link = get_object_or_404(SocialLink, id=link_id)
    else:
        link = SocialLink()
    
    if request.method == 'POST':
        link.platform = request.POST.get('platform', '')
        link.url = request.POST.get('url', '')
        link.icon = request.POST.get('icon', '')
        link.sort_order = int(request.POST.get('sort_order', 0))
        link.is_active = request.POST.get('is_active') == 'on'
        link.save()
        messages.success(request, 'Social link saved successfully.')
        return redirect('dashboard:social_links_list')
    
    context = {'link': link}
    return render(request, 'dashboard/social_link_edit.html', context)


@dashboard_required
@require_POST
def social_link_delete(request, link_id):
    """Delete a social link"""
    link = get_object_or_404(SocialLink, id=link_id)
    link.delete()
    messages.success(request, 'Social link deleted successfully.')
    return redirect('dashboard:social_links_list')


# ============================================================================
# Footer
# ============================================================================

@dashboard_required
def footer_edit(request):
    """Edit footer content"""
    footer, created = Footer.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        footer.copyright_text = request.POST.get('copyright_text', '')
        footer.additional_text = request.POST.get('additional_text', '')
        footer.save()
        messages.success(request, 'Footer updated successfully.')
        return redirect('dashboard:footer_edit')
    
    context = {'footer': footer}
    return render(request, 'dashboard/footer_edit.html', context)